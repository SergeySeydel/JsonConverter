from tkinter import *
import tkinter as tk
import xlwt
import openpyxl
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

def nullStr ( r, c ):
    if ws.cell ( row=r, column=c ).value is None:
        return "null"
    else:
        return "\"" + str ( ws.cell(row=r, column=c).value ) + "\""

def nullNum ( r, c ):
    if ws.cell ( row=r, column=c ).value is None:
        return 0
    else:
        return 1

book = xlwt.Workbook(encoding="utf-8")      # Ставит кодировку utf-8
#wb = openpyxl.load_workbook("Kib.xslx")       #Открывает файл
wb = openpyxl.load_workbook(file_path)       #Открывает файл

ws = wb.active
i = ws.max_row
print("amount of rows ")
print (i)
w = ws.max_column
print("Аmount of columns ")
print (w)

root = Tk()
root.title('Listbox')
root.geometry("800x600")
my_listbox = Listbox(root)
my_listbox.pack(pady=15)
startCols = {}

f = []

for Ci in range(5,w,1):
  #print("" + str(Ci) + str( sheet.cell(row=2, column=Ci).value ))
  if ws.cell(row=3, column=Ci).value == "Предмет":
    startCols[ws.cell(row=2, column=Ci).value] = Ci
    print( "    Группа " + str( ws.cell(row=2, column=Ci).value ))
    f = str( ws.cell(row=2, column=Ci).value)
    for Ci in range(5,w,1):
        f = str( ws.cell(row=2, column=Ci).value)
        my_list = [f]






for item in my_list:
    my_listbox.insert(END, item)


def delete():
    my_listbox.delete(ANCHOR)

def select():
    my_label.config(text=my_listbox.get(ANCHOR))
    my_label.config(text="Select", command=select)


my_button2 = Button(root, text="Select", command=select)
my_button2.pack(pady=10)

my_label = Label(root, text=' ')
my_label.pack(pady=10)

global my_label1
my_label1 = Label(root,text=' ')
my_label1.pack(pady=10)

root.mainloop()