import openpyxl  #Открывает эксель, может читать значения
#import json  #Сохраняет файл в json, сложности с форматированием
#import PySimpleGUI as sg     #Графическая оболочка, которую я подрублю после того, как реализую нормальное считывание групп
#import pandas as pd  #Нужно для форматирование файла json
import xlwt  #Поскольку возникает проблема со считыванием значения столбцов, временно поставил данную библиотеку

import tkinter as tk
from tkinter import filedialog

#import PySimpleGUI as sg
#layout = [
#    [sg.Text('File 1'), sg.InputText(), sg.FileBrowse(),
#     ],
#     ],
#    [sg.Output(size=(88, 20))],
#    [sg.Submit(), sg.Cancel()]
#]
#window = sg.Window('File Compare', layout)
#while True:                             # The Event Loop
#    event, values = window.read()
#    # print(event, values) #debug
#    if event in (None, 'Exit', 'Cancel'):
#        break

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

def nullStr ( r, c ):
    if ws.cell ( row=r, column=c ).value is None:
        return "null"
    else:
        return "\"" + str ( ws.cell(row=r, column=c).value ) + "\""

def nullNum ( r, c ):
    if ws.cell ( row=r, column=c ).value is None:
        return 0
    else:
        return 1

book = xlwt.Workbook(encoding="utf-8")      # Ставит кодировку utf-8
#wb = openpyxl.load_workbook("Kib.xslx")       #Открывает файл
wb = openpyxl.load_workbook(file_path)       #Открывает файл

ws = wb.active               #Назначает открывшийся файл эксель в переменную, чтобы потом с ним работать
i = ws.max_row               # Retrieve the maximum amount of rows
print("amount of rows ")        #Выводит количество строчек
print (i)                       #Печатает i
w = ws.max_column            # Retrieve the maximum amount of columns
print("Аmount of columns ")     #Печатает столбцы
print (w)                       #Печатает столбцы

startCols = {}

for Ci in range(5,w,1):
  #print("" + str(Ci) + str( sheet.cell(row=2, column=Ci).value ))
  if ws.cell(row=3, column=Ci).value == "Предмет":
    startCols[ws.cell(row=2, column=Ci).value] = Ci
    print( "    Группа " + str( ws.cell(row=2, column=Ci).value ) )

print('Enter name of group') # Запрашиваем название групп
j = [[], []] #Массив, в который будет записываться лекции/практики с пн по сб
counter2 = 1 #Счетчик, который пригодится потом
a = str(input()) #Запрашиваем на ввод значение(строку string), записываем его в переменную a
if a not in startCols:
  print ( "Группа '" + a + "' не существует")
  exit (0)
week = "["
c0 = startCols[a]
c1 = c0 + 1
c2 = c1 + 1
c3 = c2 + 1
for rN in range(4,74,12):
#    day = []
    for h in range(1,12,2):
        r1 = rN + h
        r2 = r1 + 1
        f1 = nullNum ( r1, c0 )
        f2 = nullNum ( r2, c0 )
        if ws.cell( row=r1, column=c0 ).value == ws.cell( row=r2, column=c0).value:
            week = week + "{" +   \
                "\"name\" : "    + nullStr( r1, c0 ) + "," + \
                "\"type\" : "    + nullStr( r1, c1 ) + "," + \
                "\"teacher\" : " + nullStr( r1, c2 ) + "," + \
                "\"room\" : "    + nullStr( r1, c3 ) + "," + \
                "\"week\" : null"  + \
            "},"
        else:
            week = week + "["
            if 0 != f1:
                week = week + "{" + \
                        "\"name\" : "    + nullStr( r1, c0 ) + "," + \
                        "\"type\" : "    + nullStr( r1, c1 ) + "," + \
                        "\"teacher\" : " + nullStr( r1, c2 ) + "," + \
                        "\"room\" : "    + nullStr( r1, c3 ) + "," + \
                        "\"week\" : 1"   + \
                    "}"
                if 0 != f2:
                    week = week + ","
            if 0 != f2:
                week = week + "{"  + \
                    "\"name\" : "    + nullStr( r2, c0 ) + "," + \
                    "\"type\" : "    + nullStr( r2, c1 ) + "," + \
                    "\"teacher\" : " + nullStr( r2, c2 ) + "," + \
                    "\"room\" : "    + nullStr( r2, c3 ) + "," + \
                    "\"week\" : 2" + \
                "}"
            week = week + "],"
#        day.append( r )
#    week.append(week)
week = week.rstrip(",") + "]"

print(week)

#with open("Data.json", "w", encoding="utf-8") as file:
#    json.dump(week, file)
#json_file = week

#with open("Data.json", "w", encoding="utf-8") as file:
#    file.writelines(str(json_file))
#    data = json.load(json_file)
#    print(data)


