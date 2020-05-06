from tkinter import *
import tkinter as tk
import xlwt
import openpyxl
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

def nullStr ( r, c ):
    if ws.cell ( row=r, column=c ).value is None:
        return "null"
    else:
        return "\"" + str ( ws.cell(row=r, column=c).value ) + "\""

def nullNum ( r, c ):
    if ws.cell ( row=r, column=c ).value is None:
        return 0
    else:
        return 1

book = xlwt.Workbook(encoding="utf-8")      # Ставит кодировку utf-8
#wb = openpyxl.load_workbook("Kib.xslx")       #Открывает файл
wb = openpyxl.load_workbook(file_path)       #Открывает файл

ws = wb.active
i = ws.max_row
print("amount of rows ")
print (i)
w = ws.max_column
print("Аmount of columns ")
print (w)

def groupToJSoN(grpID):
    week = "["
    c0 = startCols[grpID]
    c1 = c0 + 1
    c2 = c1 + 1
    c3 = c2 + 1
    for rN in range(4,74,12):
    #    day = []
        for h in range(1,12,2):
            r1 = rN + h
            r2 = r1 + 1
            f1 = nullNum ( r1, c0 )
            f2 = nullNum ( r2, c0 )
            if ws.cell( row=r1, column=c0 ).value == ws.cell( row=r2, column=c0).value:
                week = week + "{" +   \
                    "\"name\" : "    + nullStr( r1, c0 ) + "," + \
                    "\"type\" : "    + nullStr( r1, c1 ) + "," + \
                    "\"teacher\" : " + nullStr( r1, c2 ) + "," + \
                    "\"room\" : "    + nullStr( r1, c3 ) + "," + \
                    "\"week\" : null"  + \
                "},"
            else:
                week = week + "["
                if 0 != f1:
                    week = week + "{" + \
                            "\"name\" : "    + nullStr( r1, c0 ) + "," + \
                            "\"type\" : "    + nullStr( r1, c1 ) + "," + \
                            "\"teacher\" : " + nullStr( r1, c2 ) + "," + \
                            "\"room\" : "    + nullStr( r1, c3 ) + "," + \
                            "\"week\" : 1"   + \
                        "}"
                    if 0 != f2:
                        week = week + ","
                if 0 != f2:
                    week = week + "{"  + \
                        "\"name\" : "    + nullStr( r2, c0 ) + "," + \
                        "\"type\" : "    + nullStr( r2, c1 ) + "," + \
                        "\"teacher\" : " + nullStr( r2, c2 ) + "," + \
                        "\"room\" : "    + nullStr( r2, c3 ) + "," + \
                        "\"week\" : 2" + \
                    "}"
                week = week + "],"
    #        day.append( r )
    #    week.append(week)
    return week.rstrip(",") + "]"

root = Tk()
root.title('Listbox')
root.geometry("800x600")
my_listbox = Listbox(root,selectmode=SINGLE)
my_listbox.pack(pady=15)
groupNames = []
startCols = []


for Ci in range(5,w,1):
  #print("" + str(Ci) + str( sheet.cell(row=2, column=Ci).value ))
  if ws.cell(row=3, column=Ci).value == "Предмет":
    gName = str( ws.cell(row=2, column=Ci).value )
    groupNames.append ( gName )
    startCols.append ( Ci )
    my_listbox.insert ( END, gName )

#j = [[], []]
#counter2 = 1
#a = str(input())
#if a not in startCols:
#  print ( "Группа '" + a + "' не существует")
#  exit (0)

#for item in my_list:
#   my_listbox.insert(END, item)

def select():
    my_label.config(text="Select", command=select)

def saveResult():
    idx = my_listbox.curselection()[0]
    grp = groupNames[idx]
    json = groupToJSoN ( idx )
    file_path = filedialog.asksaveasfilename()
    with open ( file_path, 'w' ) as file:
        file.write ( json )
        file.close ()



my_button2 = Button(root, text="Выгрузить расписание", command=saveResult)
my_button2.pack(pady=10)

my_label = Label(root, text=' ')
my_label.pack(pady=10)

global my_label1
my_label1 = Label(root,text=' ')
my_label1.pack(pady=10)

root.mainloop()